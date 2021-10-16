# encode=utf-8
"""
Judge and query the original table data one by one,
and finally save a copy of the original data with temperature information
and a data table of heat-resistant bacteria (growth temperature greater than 50 or thermo keyword)
"""
import os.path
import sys
import time
import openpyxl
from tqdm import tqdm
from temperature import args, sqlite


def query(sqlite_db, source_file_path, target_file_path, target_file_only_thermo_path):
    if not os.path.exists(sqlite_db):
        print(f'query.py -> file not exist: {sqlite_db}')
        sys.exit(1)
    if not os.path.exists(source_file_path):
        print(f'query.py -> file not exist: {source_file_path}')
        sys.exit(1)

    # load database
    database = sqlite.Sqlite(sqlite_db)

    # save query res
    target_workbook = openpyxl.Workbook(write_only=True)
    target_sheet = target_workbook.create_sheet('sheet0')

    # only save thermo data
    thermo_workbook = openpyxl.Workbook(write_only=True)
    thermo_sheet = thermo_workbook.create_sheet('sheet0')

    # load source xlsx
    source_workbook = openpyxl.load_workbook(source_file_path, read_only=True)
    source_sheet_name = source_workbook.sheetnames[0]
    source_sheet_data = source_workbook[source_sheet_name]

    sheet_header = ['Entry', 'Entry name', 'Status', 'Protein names', 'Gene names', 'Organism', 'Length', 'Search url', 'Target url', 'Temperature']
    thermo_header = ['Entry', 'Entry name', 'Status', 'Protein names', 'Gene names', 'Organism', 'Length', 'Temperature']
    # print(sheet_header)

    # get Organism column index
    organism_col_index = sheet_header.index("Organism")

    # write header to target file
    target_sheet.append(sheet_header)
    thermo_sheet.append(thermo_header)

    # get xlsx row num
    sheet_row_num = 0
    for _ in source_sheet_data.rows:
        sheet_row_num += 1

    # read source file
    for row in tqdm(source_sheet_data.rows, desc='query processing:', total=sheet_row_num):
        strs = row[organism_col_index].value

        # delete sp.
        index = strs.find("sp.")
        if index != -1:
            strs = strs[:index]

        # delete () and data
        index = strs.find("(")
        if index != -1:
            strs = strs[:index]

        # delete [] only
        strs = strs.replace("[", "")
        strs = strs.replace("]", "")

        # keep two words
        strs_list = strs.split(" ")
        if len(strs_list) >= 2:
            strs = strs_list[0] + " " + strs_list[1]

        strs = strs.strip()

        # write source data
        data = []
        thermo_data = []
        for cell in row:
            data.append(cell.value)
            thermo_data.append(cell.value)

        # check thermo
        if "thermo" in strs or "Thermo" in strs:
            data.append("")
            data.append("")
            data.append("thermo")
            target_sheet.append(data)

            thermo_data.append("thermo")
            thermo_sheet.append(thermo_data)

        else:
            # query database
            result = database.select(strs)

            # is thermo ?
            is_thermo = False

            # write query result
            if len(result) > 0:
                data.append(result[0][2])
                data.append(result[0][3])
                data.append(result[0][4])

                if result[0][4] != "None":
                    if int(result[0][4]) >= 50:
                        thermo_data.append(result[0][4])
                        is_thermo = True

            else:
                print(strs)

            # write to sheet
            if is_thermo:
                thermo_sheet.append(thermo_data)

            target_sheet.append(data)

    # save file
    if os.path.exists(target_file_only_thermo_path):
        os.remove(target_file_only_thermo_path)
    thermo_workbook.save(target_file_only_thermo_path)

    if os.path.exists(target_file_path):
        os.remove(target_file_path)
    target_workbook.save(target_file_path)

    # close file
    source_workbook.close()

    time.sleep(1)


if __name__ == '__main__':
    args = args.Args()

    query(args.sqlite_bacterium_temperature_db, args.source_xlsx_file, args.query_source_no_thermo_xlsx_result_file, args.all_result_xlsx_file)
