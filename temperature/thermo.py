# encoding: utf-8
"""
author: yudd
time  : 2021.9.30
use   : It is divided into two files according to whether the organization column of the source file contains the thermo keyword
> https://www.osgeo.cn/openpyxl/optimized.html#read-only-mode
"""

import os
import sys
import time
import openpyxl
from tqdm import tqdm
from temperature import args


def find_thermo(file, is_thermo_file, no_thermo_file):
    """
    When the cell of the Organism column contains thermo characters, \n
    the row will be written to the file_is_thermo file, \n
    otherwise it will be written to the file_no_thermo file. \n
    :param file: xlsx file absolute path
    :param is_thermo_file: new excel for have thermo key
    :param no_thermo_file: new execl for no thermo key
    :return: none
    """
    # check xlsx file
    if os.path.exists(file):
        print(f'thermo.py using -> {file}')
    else:
        print(f'{file} is not exist')
        sys.exit(1)

    # read sheet
    ld = openpyxl.load_workbook(file, read_only=True)  # openpyxl begin [1, 1] and set read only mode
    sheet_name = ld.sheetnames[0]  # get the first sheet by default, and sheet type is list
    sheet_data = ld[sheet_name]    # get data
    # sheet = ls["sheet0"]

    # get header
    sheet_header = []
    for header in sheet_data.iter_rows(min_row=1, min_col=1, max_row=1):
        # head is every col cell data
        for item in header:
            sheet_header.append(item.value)
    print(f'header: {sheet_header}')

    # get sheet row
    sheet_row_num = 0
    for _ in sheet_data.rows:
        sheet_row_num += 1

    # get Organism column index
    organism_col_index = sheet_header.index("Organism")
    print(f'organism index: {organism_col_index}')

    # create write object and write header
    is_thermo_wb = openpyxl.Workbook(write_only=True)
    is_thermo_ws = is_thermo_wb.create_sheet("sheet0")
    is_thermo_ws.append(sheet_header)

    no_thermo_wb = openpyxl.Workbook(write_only=True)
    no_thermo_ws = no_thermo_wb.create_sheet("sheet0")
    # no_thermo_ws.append(sheet_header)

    #
    for row in tqdm(sheet_data.rows, desc='thermo processing:', total=sheet_row_num):
        strs = row[organism_col_index].value
        data = []
        for cell in row:
            data.append(cell.value)

        if "thermo" in strs or "Thermo" in strs:
            is_thermo_ws.append(data)
        else:
            no_thermo_ws.append(data)

    # delete new file first
    if os.path.exists(is_thermo_file):
        os.remove(is_thermo_file)
    if os.path.exists(no_thermo_file):
        os.remove(no_thermo_file)

    # save file
    is_thermo_wb.save(is_thermo_file)
    no_thermo_wb.save(no_thermo_file)
    print(f'thermo keyword data save to {is_thermo_file}')
    print(f'no-thermo keyword data save to {no_thermo_file}\n')

    # ues read only mode, and need close file
    ld.close()

    time.sleep(1)


if __name__ == '__main__':
    args = args.Args()

    # param: source excel file; save thermo data xlsx; save no thermo data xlsx
    find_thermo(args.source_xlsx_file, args.is_thermo_source_xlsx_file, args.no_thermo_source_xlsx_file)
