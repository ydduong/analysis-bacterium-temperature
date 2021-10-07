# encode=utf-8
import os.path

import openpyxl
from tqdm import tqdm
from analysis import main, sqlite

args = main.Args()
target_file_path = args.query_result_temperature
source_file_path = args.source_file
target_file_only_thermo_path = args.query_result_only_thermo
print(f'source file: {source_file_path}')
print(f'target file: {target_file_path}')
print(f'only thermo file: {target_file_only_thermo_path}')

# load database
database = sqlite.Sqlite(args.sqlite)

# word = "Candidatus Wildermuthbacteria"
# res = database.select(word)
# print(res)

# save query res
target_workbook = openpyxl.Workbook(write_only=True)
target_sheet = target_workbook.create_sheet('sheet0')

# only save thermo data
thermo_workbook = openpyxl.Workbook(write_only=True)
thermo_sheet = thermo_workbook.create_sheet('sheet0')

# load source xlsx
source_workbook = openpyxl.load_workbook(source_file_path, read_only=True)
source_sheet_name = source_workbook.sheetnames[0]
source_sheet_data = source_workbook[source_sheet_name]

# get source xlsx header
# sheet_header = []
# for header in source_sheet_data.iter_rows(min_row=1, min_col=1, max_row=1):
#     # head is every col cell data
#     for item in header:
#         sheet_header.append(item.value)
#
# # add new header
# sheet_header.append('Search url')
# sheet_header.append('Target url')
# sheet_header.append('Temperature')
sheet_header = ['Entry', 'Entry name', 'Status', 'Protein names', 'Gene names', 'Organism', 'Length', 'Search url', 'Target url', 'Temperature']
thermo_header = ['Entry', 'Entry name', 'Status', 'Protein names', 'Gene names', 'Organism', 'Length', 'Temperature']
print(sheet_header)

# get Organism column index
organism_col_index = sheet_header.index("Organism")

# write header to target file
target_sheet.append(sheet_header)
thermo_sheet.append(thermo_header)

# get xlsx row num
sheet_row_num = 0
for _ in source_sheet_data.rows:
    sheet_row_num += 1

# read source file
for row in tqdm(source_sheet_data.rows, desc='processing:', total=sheet_row_num):
    strs = row[organism_col_index].value

    # delete sp.
    index = strs.find("sp.")
    if index != -1:
        strs = strs[:index]

    # delete () and data
    index = strs.find("(")
    if index != -1:
        strs = strs[:index]

    # delete [] only
    strs = strs.replace("[", "")
    strs = strs.replace("]", "")

    # keep two words
    strs_list = strs.split(" ")
    if len(strs_list) >= 2:
        strs = strs_list[0] + " " + strs_list[1]

    # write source data
    data = []
    thermo_data = []
    for cell in row:
        data.append(cell.value)
        thermo_data.append(cell.value)

    # check thermo
    if "thermo" in strs or "Thermo" in strs:
        data.append("")
        data.append("")
        data.append("thermo")
        target_sheet.append(data)

        thermo_data.append("thermo")
        thermo_sheet.append(thermo_data)

    else:
        # query database
        result = database.select(strs)

        # is thermo ?
        is_thermo = False

        # write query result
        if len(result) > 0:
            data.append(result[0][2])
            data.append(result[0][3])
            data.append(result[0][4])

            if result[0][4] != "None":
                if int(result[0][4]) >= 50:
                    thermo_data.append(result[0][4])
                    is_thermo = True

        else:
            print(strs)

        # write to sheet
        if is_thermo:
            thermo_sheet.append(thermo_data)

        target_sheet.append(data)

# save file
if os.path.exists(target_file_only_thermo_path):
    os.remove(target_file_only_thermo_path)
thermo_workbook.save(target_file_only_thermo_path)

if os.path.exists(target_file_path):
    os.remove(target_file_path)
target_workbook.save(target_file_path)

# close file
source_workbook.close()
