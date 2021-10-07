# encoding: utf-8
"""
author: yudd
time  : 2021.10.2
use   : data preprocessing organism name to sqlite
"""
import os
import openpyxl
from analysis import main, thermo
from tqdm import tqdm


def preprocessing(file, database_xlsx):
    """
    Keep only the first two words and delete sp. ...
    :param file: excel absolute path
    :param database_xlsx: save results to this file
    :return:
    """
    # check xlsx file
    if os.path.exists(file):
        print(f'using: {file}')
    else:
        print(f'{file} is not exist')

    # read sheet
    ld = openpyxl.load_workbook(file, read_only=True)  # openpyxl begin [1, 1] and set read only mode
    sheet_name = ld.sheetnames[0]  # get the first sheet by default, and sheet type is list
    sheet_data = ld[sheet_name]    # get data

    # get header
    sheet_header = []
    for header in sheet_data.iter_rows(min_row=1, min_col=1, max_row=1):
        # head is every col cell data
        for item in header:
            sheet_header.append(item.value)
    # print(f'header: {sheet_header}')

    # get Organism column index
    organism_col_index = sheet_header.index("Organism")
    # print(f'organism index: {organism_col_index}')

    # create write object and write header
    database_wb = openpyxl.Workbook(write_only=True)
    database_ws = database_wb.create_sheet("sheet0")

    # find thermo str
    no_thermo_set = set()
    for row in tqdm(sheet_data.rows, desc='processing:', total=sheet_data.max_row):
        strs = row[organism_col_index].value

        # delete sp.
        index = strs.find("sp.")
        if index != -1:
            strs = strs[:index]

        # delete () and data
        index = strs.find("(")
        if index != -1:
            strs = strs[:index]

        # delete [] only
        strs = strs.replace("[", "")
        strs = strs.replace("]", "")

        # keep two words
        strs_list = strs.split(" ")
        if len(strs_list) >= 2:
            strs = strs_list[0] + " " + strs_list[1]

        # duplicate removal
        no_thermo_set.add(strs)

    # save to sheet
    for item in no_thermo_set:
        database_ws.append([item])

    # delete new file first
    if os.path.exists(database_xlsx):
        os.remove(database_xlsx)

    # save file
    database_wb.save(database_xlsx)

    # ues read only mode, and need close file
    ld.close()


if __name__ == '__main__':
    args = main.Args()

    preprocessing(args.no_thermo_xlsx, args.database)




