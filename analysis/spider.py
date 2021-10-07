# encoding: utf-8
"""
author: yudd
time  : 2021.9.30
use   : to read excel file and find thermo key word
12000 * 1.0 / 60.0 / 60.0

还有两个问题：
  程序不能停止
  部分链接会多一个空格
"""

import os
import sys
import queue
import time
import openpyxl
import urllib3
import requests
import threading
from lxml import etree
from tqdm import tqdm
from analysis import main


class Spider:
    def __init__(self, file, result, log):
        """
        :param file: excel absolute path
        :param result: save result file
        :param log: to print log
        """
        if not os.path.exists(file):
            print(f'{file} is not exist.')
            sys.exit(1)

        self._file = file
        self._file_load_book = None
        self._tqdm = None

        # save file
        self._save_file = result
        if os.path.exists(self._save_file):
            os.remove(self._save_file)
        # self._workbook = openpyxl.Workbook(write_only=True)
        # self._sheet = self._workbook.create_sheet("sheet0")

        self.logger = log
        self.logger.append('self.__init__()')

        self._base_url = "https://webshop.dsmz.de/index.php?lang=1&cl=search&searchparam="
        self._header = {
            'sec-ch-ua': '"Google Chrome";v="93", " Not;A Brand";v="99", "Chromium";v="93"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"',
            'Upgrade-Insecure-Requests': '1',
            'Connection': 'close',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36',
        }
        self._header2 = {
            'Connection': 'close'
        }

        self._start_url_q = queue.Queue()
        self._target_url_q = queue.Queue()
        self._target_info_q = queue.Queue()
        self._thread_lock = threading.Lock()

    def get_start_url(self):
        """
        read xlsx file and create start url, save to self._start_url_q
        :return: None
        """
        self.logger.append('self.get_start_url()')
        # read sheet
        self._file_load_book = openpyxl.load_workbook(self._file, read_only=True)
        _sheet_name = self._file_load_book.sheetnames[0]
        _sheet_data = self._file_load_book[_sheet_name]

        # get xlsx row num
        _sheet_row_num = 1
        for _ in _sheet_data.rows:
            _sheet_row_num += 1

        # set tqdm bar
        self._tqdm = tqdm(total=_sheet_row_num, desc='processing')

        # create url
        for row in _sheet_data.rows:
            for cell in row:
                _strs_total = cell.value
                _strs_list = _strs_total.strip().split(" ")
                _url = self._base_url
                _url += _strs_list[0]
                if len(_strs_list) > 1:
                    _url += '+' + _strs_list[1]

                # print(f'get_start_url() url: {_url}')
                # [key word, start url]
                self._start_url_q.put([_strs_total, _url])

    def get_html(self, _url):
        """
        call url and return html text
        :return:
        """
        self.logger.append('self.get_html()')
        try:
            # print(f'get_html() url: {_url}')
            _resp = requests.get(url=_url, headers=self._header2, timeout=(6, 6), verify=False)  # 发出请求
            if _resp.status_code == 200:
                self.logger.append('200 ok')
                return _resp
        except Exception as e:
            self.logger.append("\nurl: {0}) \nerror: {1}".format(_url, e))
            self.logger.append('none ok')
            return "error"

    def visit_start_url(self):
        """
        get start url from self._start_url_q, and visit it to get start html  \n
        from start html get target url and save it to self._target_url_q  \n
        and this function always execute  \n
        :return:
        """
        while True:
            self.logger.append('self.visit_start_url()')
            _key_url = self._start_url_q.get()
            self.logger.append(f'visit_start_url() start url: {_key_url[1]}')
            _resp = self.get_html(_key_url[1])
            _target_url = None
            if _resp is not None and _resp != 'error':
                _html_etree = etree.HTML(_resp.text)
                # //*[@id="cnblogs_post_body"]/p[16]
                _search_result = _html_etree.xpath('//*[@id="searchList"]/div/div/form/div[2]/div[2]/div/div[6]/div/a/@href')

                if len(_search_result) > 0:
                    _target_url = _search_result[0]

                _resp.close()

            self.logger.append(f'here here target_url: {_target_url}')

            # [key word, start url, target url]
            if _resp == 'error':
                self._start_url_q.put(_key_url)
            else:
                self._target_url_q.put([_key_url[0], _key_url[1], _target_url])

            self._start_url_q.task_done()
            time.sleep(1)

            # self._thread_lock.acquire()
            # print(self._start_url_q.qsize())
            # self._thread_lock.release()

    def visit_target_url(self):
        """
        get target url from self._target_url_q, and visit it to get target html  \n
        from target html get target info and save it to self._target_info_q  \n
        and this function always execute  \n
        :return:
        """
        while True:
            self.logger.append('self.visit_target_url()')
            # [key, start url, target url]
            _key_url = self._target_url_q.get()
            self.logger.append(f'target url: {_key_url[2]}')

            _target_info = None
            _url_443 = False
            if _key_url[2] is not None:
                _resp = self.get_html(_key_url[2])

                if _resp == 'error':
                    _url_443 = True

                if _resp is not None and _resp != 'error':
                    _html_text = _resp.text
                    _info_index = _html_text.find('°C')
                    # °C
                    if _info_index == -1:
                        _info_index = _html_text.find('&deg;C')

                    if _info_index != -1:
                        if _html_text[_info_index-2] == " ":
                            _target_info = _html_text[_info_index-2:_info_index].strip()
                        else:
                            _target_info = _html_text[_info_index-3:_info_index].strip()

                    _resp.close()

            if _url_443:
                self._target_url_q.put(_key_url)
            else:
                # [key word, start url, target url, target info]
                self._target_info_q.put([_key_url[0], _key_url[1], _key_url[2], _target_info])

            self._target_url_q.task_done()
            time.sleep(1)

    def save_info(self):
        """
        form self._target_info_q get info, and save it to self._save_file
        and this function always execute  \n
        :return:
        """
        while True:
            self.logger.append('self.save_info()')
            self._thread_lock.acquire()
            _info = self._target_info_q.get()
            with open(self._save_file, "a") as output:
                _strs = ""
                for _item in _info:
                    _strs += str(_item) + ','
                _strs += '\n'
                output.write(_strs)
            self._tqdm.update(1)
            self._target_info_q.task_done()
            self._thread_lock.release()

    def run(self):
        _thread_list = [threading.Thread(target=self.get_start_url)]
        _thread_num = [8, 8, 10]
        _thread_function = [self.visit_start_url, self.visit_target_url, self.save_info]

        for _index in range(3):
            for _ in range(_thread_num[_index]):
                _thread_list.append(threading.Thread(target=_thread_function[_index]))

        self.logger.append('self.run()')
        # _thread_list = [
        #     threading.Thread(target=self.get_start_url),
        #     threading.Thread(target=self.visit_start_url),
        #     threading.Thread(target=self.visit_target_url),
        #     threading.Thread(target=self.save_info)
        # ]

        self.logger.append(f'thread num: {len(_thread_list)}')

        for _item in _thread_list:
            _item.setDaemon(True)
            _item.start()

        for _item in _thread_list:
            _item.join()

        # for _item in [self._start_url_q, self._target_url_q, self._target_info_q]:
        #     _item.join()

        # time.sleep(100)

        # save and close file
        # self._workbook.save(self._save_file)
        self._tqdm.close()
        self._file_load_book.close()

        print("ok")


if __name__ == '__main__':
    urllib3.disable_warnings()

    args = main.Args()
    logger = main.Log()
    logger.is_print(False)

    # create spider
    print(f'xlsx: {args.database}, result: {args.temperature}')
    spider = Spider(args.database2, args.temperature2, logger)
    spider.run()
