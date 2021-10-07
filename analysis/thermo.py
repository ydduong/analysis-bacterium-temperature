# encoding: utf-8
"""
author: yudd
time  : 2021.9.30
use   : to read excel file and find thermo key word
> https://www.osgeo.cn/openpyxl/optimized.html#read-only-mode
"""

import os
import time
import openpyxl
from tqdm import tqdm
from analysis import main


def find_thermo(file, is_thermo_file, no_thermo_file):
    """
    When the cell of the Organism column contains thermo characters, \n
    the row will be written to the file_is_thermo file, \n
    otherwise it will be written to the file_no_thermo file. \n
    :param file: xlsx file absolute path
    :param is_thermo_file: new excel for have thermo key
    :param no_thermo_file: new execl for no thermo key
    :return: none
    """

    # time
    start_time = time.time()

    # check xlsx file
    if os.path.exists(file):
        print(f'using: {file}')
    else:
        print(f'{file} is not exist')

    # read sheet
    ld = openpyxl.load_workbook(file, read_only=True)  # openpyxl begin [1, 1] and set read only mode
    sheet_name = ld.sheetnames[0]  # get the first sheet by default, and sheet type is list
    sheet_data = ld[sheet_name]    # get data
    # sheet = ls["sheet0"]

    # get header
    sheet_header = []
    for header in sheet_data.iter_rows(min_row=1, min_col=1, max_row=1):
        # head is every col cell data
        for item in header:
            sheet_header.append(item.value)
    print(f'header: {sheet_header}')

    # get Organism column index
    organism_col_index = sheet_header.index("Organism")
    print(f'organism index: {organism_col_index}')

    # test: print the first 100 rows of data in organism column
    # num = 0
    # for row in sheet_data.rows:
    #     print(row[organism_col_index].value)
    #     num += 1
    #     if num == 100:
    #         break

    # create write object and write header
    is_thermo_wb = openpyxl.Workbook(write_only=True)
    is_thermo_ws = is_thermo_wb.create_sheet("sheet0")
    is_thermo_ws.append(sheet_header)

    no_thermo_wb = openpyxl.Workbook(write_only=True)
    no_thermo_ws = no_thermo_wb.create_sheet("sheet0")
    # no_thermo_ws.append(sheet_header)

    # find thermo str
    # for row in sheet_data.rows:
    #     pass

    for row in tqdm(sheet_data.rows, desc='processing:', total=sheet_data.max_row):
        strs = row[organism_col_index].value
        data = []
        for cell in row:
            data.append(cell.value)

        if "thermo" in strs or "Thermo" in strs:
            is_thermo_ws.append(data)
        else:
            no_thermo_ws.append(data)

    # delete new file first
    if os.path.exists(is_thermo_file):
        os.remove(is_thermo_file)
    if os.path.exists(no_thermo_file):
        os.remove(no_thermo_file)

    # save file
    is_thermo_wb.save(is_thermo_file)
    no_thermo_wb.save(no_thermo_file)

    # ues read only mode, and need close file
    ld.close()

    print(f'use time is {time.time() - start_time}')


if __name__ == '__main__':
    args = main.Args()

    find_thermo(args.xlsx_file, args.is_thermo_xlsx, args.no_thermo_xlsx)
