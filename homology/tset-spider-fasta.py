# encode=utf-8

from homology.args import Args
from tqdm import tqdm
import requests
import openpyxl


# get path
args = Args()

# source excel file
homology_file = args.homology_file

# target data save dir
fasta_data_dir = args.fasta_dir_path


# get html
def get_html(_url,):
    try:
        _resp = requests.get(url=_url, headers={'Connection': 'close'}, timeout=(6., 6))  # 发出请求
        _resp.encoding = _resp.apparent_encoding
        if _resp.status_code == 200:
            return _resp
    except Exception as e:
        print("地址({0})出错：{1}".format(_url, e))


url_prefix = 'https://www.uniprot.org/uniprot/'
urls = list()
# url = [
#     'https://www.uniprot.org/uniprot/A0A088SZA7.fasta',
# ]

# read source excel file
workbook = openpyxl.load_workbook(homology_file, read_only=True)
sheet_name = workbook.sheetnames[0]
sheet = workbook[sheet_name]

# set up url
for entry_number in sheet.iter_rows(min_row=1, min_col=1, max_col=1):
    cell = entry_number[0]
    entry = cell.value
    urls.append([entry, url_prefix+entry+'.fasta'])
    # print(cell.value)

workbook.close()

# delete header row
urls.pop(0)

# get data form web
for url in tqdm(urls, total=len(urls), desc='processing:'):
    resp = get_html(url[1])

    # write
    file = fasta_data_dir + f'/{url[0]}.txt'
    with open(file, 'w') as f:
        f.write(resp.text)






