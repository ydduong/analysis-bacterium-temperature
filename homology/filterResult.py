# encode=utf-8
import os.path

from homology.args import Args
import openpyxl


def filter_result(args):
    # source excel file
    homology = args.homology_file
    # source txt file
    result_homology = args.result_homology_file
    # target excel file
    result_non_homology = args.result_non_homology_file
    if os.path.exists(result_non_homology):
        os.remove(result_non_homology)

    target_workbook = openpyxl.Workbook(write_only=True)
    target_sheet = target_workbook.create_sheet('sheet0')

    # print(homology, result_homology, result_non_homology)

    # read source txt file
    homology_list = list()
    with open(result_homology, 'r', encoding='utf-8') as reader:
        lines = reader.readlines()
        for line in lines:
            homology_list.append(line.strip())

    # read source excel file
    source_wordbook = openpyxl.load_workbook(homology, read_only=True)
    source_sheet_first_name = source_wordbook.sheetnames[0]
    source_sheet = source_wordbook[source_sheet_first_name]

    # duplicate removal for source excel file
    entry_set = set()
    entry_set_num = 0
    for row in source_sheet.rows:
        first_col = row[0].value

        entry_set.add(first_col)
        entry_set_num += 1
        if entry_set_num != len(entry_set):
            entry_set_num -= 1
            continue

        row_data = list()
        for cell in row:
            row_data.append(cell.value)

        if first_col in homology_list:
            target_sheet.append(row_data)

    source_wordbook.close()
    target_workbook.save(result_non_homology)


if __name__ == '__main__':
    filter_result(Args())


