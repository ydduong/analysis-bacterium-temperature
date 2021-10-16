# encode=utf-8
import queue
import threading
import os
import shutil
from homology.args import Args
from tqdm import tqdm
import requests
import openpyxl
from homology.preprocessingFasta import merge_fasta


# get path
args = Args()

# source excel file
homology_file = args.homology_file

# target data save dir
fasta_data_dir = args.fasta_dir_path

is_delete_fasta = True
if is_delete_fasta:
    if os.path.exists(fasta_data_dir):
        shutil.rmtree(fasta_data_dir)

    if not os.path.exists(fasta_data_dir):
        os.makedirs(fasta_data_dir)

url_prefix = 'https://www.uniprot.org/uniprot/'
urls = list()
urls_set = set()

# read source excel file
workbook = openpyxl.load_workbook(homology_file, read_only=True)
sheet_name = workbook.sheetnames[0]
sheet = workbook[sheet_name]

# set up url
for entry_number in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
    cell = entry_number[0]
    entry = cell.value
    urls_set.add(entry)

for url in urls_set:
    urls.append([url, url_prefix+url+'.fasta'])
    # print(cell.value)

# print(urls)
workbook.close()


# threading
class Spider:
    def __init__(self, _urls, _file_save_dir):
        self._urls = _urls
        self._file_save_dir = _file_save_dir
        self._headers = {'Connection': 'close'}

        self._queue_urls = queue.Queue()
        self._queue_htmls = queue.Queue()
        self._tqdm = tqdm(desc='processing', total=len(self._urls))

    def get_html(self, _url):
        try:
            _resp = requests.get(url=_url, headers=self._headers, timeout=(6., 6))  # 发出请求
            _resp.encoding = _resp.apparent_encoding
            if _resp.status_code == 200:
                return _resp
            else:
                print(f'url: {_url}, status code: {_resp.status_code}')
                return 'error code'
        except Exception as e:
            print(f'url: {url}, error：{e}')
            return None

    def set_url(self):
        for _url in self._urls:
            # [key, url]
            self._queue_urls.put(_url)

    def visit_fasta_html(self):
        while True:
            _key_url = self._queue_urls.get()
            _resp = self.get_html(_key_url[1])
            _html_text = ""
            if _resp is not None:
                if _resp != 'error code':
                    _html_text = _resp.text
                    # key html
                    self._queue_htmls.put([_key_url[0], _html_text])
                else:
                    self._tqdm.update()

            else:
                self._queue_urls.put(_key_url)

            self._queue_urls.task_done()

    def save_info(self):
        while True:
            _key_html = self._queue_htmls.get()
            _file = self._file_save_dir + f'/{_key_html[0]}.txt'
            with open(_file, 'w') as _f:
                _f.write(_key_html[1])
            self._tqdm.update()
            self._queue_htmls.task_done()

    def run(self):
        _thread_list = [threading.Thread(target=self.set_url)]
        _thread_num = [8, 8]
        _thread_function = [self.visit_fasta_html, self.save_info]

        for _index in range(2):
            for _ in range(_thread_num[_index]):
                _thread_list.append(threading.Thread(target=_thread_function[_index]))

        for _thread in _thread_list:
            _thread.setDaemon(True)
            _thread.start()

        for _queue in [self._queue_urls, self._queue_htmls]:
            _queue.join()


spider = Spider(urls, fasta_data_dir)
spider.run()

# merge file to all-fasta.txt
merge_fasta(args)
